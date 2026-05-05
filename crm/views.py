from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
import pandas as pd
import numpy as np
from .models import Deal, Customer, Contact
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.contrib import messages
import re
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse
from django.contrib.humanize.templatetags.humanize import intcomma
from reportlab.platypus import Paragraph

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('/dashboard/')  # or your home page

        else:
            messages.error(request, "Invalid username or password")
            return redirect('login')

    return render(request, 'login.html')


def register_view(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        confirm_password = request.POST.get('confirm', '')

        print('test123',username,email,password,confirm_password)

        # Username exists
        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already exists")
            print("user check")
            return redirect('register')

        # ✅ Create user
        User.objects.create_user(
            username=username,
            email=email,
            password=password
        )

        return redirect('login')

    return render(request, 'register.html')


def dashboard(request):
    contacts = Contact.objects.filter(owner=request.user).values()
    all_contacts = Contact.objects.all().values()
    deals = Deal.objects.filter(owner=request.user).values()
    all_deals = Deal.objects.all().values()
    customers = Customer.objects.all().values()

    my_contacts_count = Contact.objects.filter(owner=request.user).count()
    all_contacts_count = Contact.objects.count()

    my_deals_count = Deal.objects.filter(owner=request.user).count()
    all_deals_count = Deal.objects.count()

    # Convert to DataFrame
    df_contacts = pd.DataFrame(contacts)
    df_deals = pd.DataFrame(deals)
    df_customers = pd.DataFrame(customers)

    total_contacts = len(df_contacts)
    total_deals = len(df_deals)

    if not df_customers.empty:
        hot = df_customers[df_customers['lead_status'] == 'Hot'].shape[0]
        warm = df_customers[df_customers['lead_status'] == 'Warm'].shape[0]
        cold = df_customers[df_customers['lead_status'] == 'Cold'].shape[0]

        # NumPy example
        avg_leads = np.mean([hot, warm, cold])
    else:
        hot = warm = cold = avg_leads = 0

    #deals_qs = Deal.objects.filter(owner=request.user)
    deals_qs = Deal.objects.all()

    deal_names = []
    deal_counts = []

    for d in deals_qs:
        deal_names.append(d.title)
        deal_counts.append(
            Customer.objects.filter(deal=d).count()
        )

    return render(request, 'dashboard.html', {
        'contacts': total_contacts,
        'deals': total_deals,
        'hot': hot,
        'warm': warm,
        'cold': cold,
        'avg': avg_leads,
        'deal_names': deal_names,
        'deal_counts': deal_counts,
        'all_contacts': all_contacts,
        'all_deals': all_deals,
        'my_contacts_count': my_contacts_count,
        'all_contacts_count': all_contacts_count,
        'my_deals_count': my_deals_count,
        'all_deals_count': all_deals_count,
    })


def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def contact_list(request):

    view_type = request.GET.get('view', 'my')  # default = my
    query = request.GET.get('q')

    if view_type == 'all':
        contacts = Contact.objects.all()
    else:
        contacts = Contact.objects.filter(owner=request.user)

    if query:
        contacts = contacts.filter(
            Q(name__icontains=query) |
            Q(email__icontains=query)
        )

    return render(request, 'contacts/list.html', {
        'contacts': contacts,
        'view_type': view_type
    })


@login_required
def add_contact(request):

    view = request.GET.get('view', 'my')
    q = request.GET.get('q', '')

    if request.method == 'POST':
        Contact.objects.create(
            name=request.POST['name'],
            email=request.POST['email'],
            phone=request.POST['phone'],
            company=request.POST['company'],
            job_title=request.POST['job_title'],
            city=request.POST['city'],
            notes=request.POST['notes'],
            owner=request.user
        )

        return redirect(f'/contacts/?view={view}&q={q}')

    return render(request, 'contacts/add.html', {
        'view': view,
        'q': q
    })


@login_required
def edit_contact(request, id):
    contact = Contact.objects.get(id=id)

    view = request.GET.get('view', 'my')
    q = request.GET.get('q', '')

    if request.method == 'POST':
        contact.name = request.POST['name']
        contact.email = request.POST['email']
        contact.phone = request.POST['phone']
        contact.company = request.POST['company']
        contact.job_title = request.POST['job_title']
        contact.city = request.POST['city']
        contact.notes = request.POST['notes']
        contact.save()

        return redirect(f'/contacts/?view={view}&q={q}')

    return render(request, 'contacts/edit.html', {
        'contact': contact,
        'view': view,
        'q': q
    })


@login_required
def delete_contact(request, id):
    contact = Contact.objects.get(id=id)

    view = request.GET.get('view', 'my')
    q = request.GET.get('q', '')

    contact.delete()

    return redirect(f'/contacts/?view={view}&q={q}')

@login_required
def contact_view(request, id):
    contact = Contact.objects.get(id=id)
    return render(request, 'contacts/view.html', {'contact': contact})

@login_required
def download_contact_pdf(request, id):
    contact = Contact.objects.get(id=id)

    response = HttpResponse(content_type='application/pdf')

    safe_name = contact.name.replace(" ", "_")
    response['Content-Disposition'] = f'attachment; filename="{safe_name}.pdf"'

    #response['Content-Disposition'] = f'attachment; filename="contact_{contact.id}.pdf"'

    doc = SimpleDocTemplate(response)
    styles = getSampleStyleSheet()

    elements = []

    # 🔥 Title
    elements.append(Paragraph("<b>Contact Details</b>", styles['Title']))
    elements.append(Spacer(1, 15))

    # Helper for empty values
    def val(v):
        return v if v else "No data specified"

    # 📦 Table data (Label | Value)
    data = [
        ["Name", val(contact.name)],
        ["Email", val(contact.email)],
        ["Phone", val(contact.phone)],
        ["Company", val(contact.company)],
        ["Job Title", val(contact.job_title)],
        ["City", val(contact.city)],
        ["Notes", Paragraph(val(contact.notes), styles['Normal'])],
    ]

    table = Table(data, colWidths=[120, 300])

    # 🎨 Styling (this creates the "card look")
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),   # left column bg
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),

        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),        # borders
        ('BOX', (0, 0), (-1, -1), 1, colors.black),

        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),     # labels bold
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),

        ('PADDING', (0, 0), (-1, -1), 10),                   # spacing
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    elements.append(table)

    elements.append(Spacer(1, 20))

    # Footer
    elements.append(Paragraph("Generated from Micro#CRM System", styles['Italic']))

    doc.build(elements)

    return response

@login_required
def download_deal_pdf(request, id):
    deal = Deal.objects.get(id=id)
    customers = deal.customer_set.all()  # adjust if related_name exists

    response = HttpResponse(content_type='application/pdf')
    #response['Content-Disposition'] = f'attachment; filename="deal_{deal.id}.pdf"'
    safe_title = deal.title.replace(" ", "_")
    response['Content-Disposition'] = f'attachment; filename="{safe_title}.pdf"'

    doc = SimpleDocTemplate(response)
    styles = getSampleStyleSheet()

    elements = []

    # 🔥 Title
    elements.append(Paragraph("<b>Deal Report</b>", styles['Title']))
    elements.append(Spacer(1, 15))

    def val(v):
        return v if v else "No data specified"

    # 📦 Deal Details (Card Style)
    deal_data = [
        ["Title", val(deal.title)],
        ["Description", Paragraph(val(deal.description), styles['Normal'])],
        ["Value", f"Rs.{intcomma(deal.value)}"],
        ["Status", val(deal.status)],
        ["Expected Close Date",
         deal.expected_close_date.strftime("%d/%m/%Y") if deal.expected_close_date else "No data specified"
        ],
    ]

    deal_table = Table(deal_data, colWidths=[160, 300])

    deal_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('PADDING', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    elements.append(deal_table)
    elements.append(Spacer(1, 20))

    # 👥 Customers Section
    elements.append(Paragraph("<b>Customers</b>", styles['Heading2']))
    elements.append(Spacer(1, 10))

    if customers.exists():
        customer_data = [["Name", "Email", "Lead"]]

        for c in customers:
            customer_data.append([
                val(c.contact.name),
                val(c.contact.email),
                val(c.lead_status)
            ])

        customer_table = Table(customer_data, colWidths=[150, 200, 100])

        customer_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))

        elements.append(customer_table)
    else:
        elements.append(Paragraph("No customers added", styles['Italic']))

    elements.append(Spacer(1, 20))

    # Footer
    elements.append(Paragraph("Generated from Micro#CRM System", styles['Italic']))

    doc.build(elements)

    return response

@login_required
def download_contacts_excel(request):
    view_type = request.GET.get('view', 'my')  # my / all

    # Filter data
    if view_type == 'all':
        contacts = Contact.objects.all()
        filename = "All_Contacts_List.xlsx"
        include_owner = True
    else:
        contacts = Contact.objects.filter(owner=request.user)
        filename = "My_Contacts_List.xlsx"
        include_owner = False

    # Prepare data
    data = []
    for c in contacts:
        row = {
            "Name": c.name or "",
            "Email": c.email or "",
            "Phone": c.phone or "",
            "Company": c.company or "",
            "Job Title": c.job_title or "",
            "City": c.city or "",
            "Notes": c.notes or "",
            "Created At": c.created_at.strftime("%d/%m/%Y") if c.created_at else "",
        }

        if include_owner:
            row["Owner"] = c.owner.username if c.owner else ""

        data.append(row)

    df = pd.DataFrame(data)

    # Create response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Write Excel
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Contacts')

        workbook = writer.book
        worksheet = writer.sheets['Contacts']

        # 🎨 Header style
        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_font = Font(bold=True)
        align_center = Alignment(vertical='center')

        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

        # Auto column width
        for col in worksheet.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            worksheet.column_dimensions[col_letter].width = max_length + 3

    return response

@login_required
def deal_list(request):
    deals = Deal.objects.filter(owner=request.user)

    deal_data = []

    for deal in deals:
        customers = Customer.objects.filter(deal=deal)

        deal_data.append({
            'deal': deal,
            'customers': customers
        })

    return render(request, 'deals/list.html', {
        'deal_data': deal_data
    })

@login_required
def other_deals(request):
    deals = Deal.objects.exclude(owner=request.user)
    return render(request, 'deals/other_list.html', {
        'deals': deals
    })

@login_required
def add_deal(request):
    if request.method == 'POST':
        Deal.objects.create(
            title=request.POST['title'],
            description=request.POST['description'],
            value=request.POST['value'],
            expected_close_date=request.POST['expected_close_date'],
            owner=request.user
        )
        return redirect('deal_list')

    return render(request, 'deals/add.html')


@login_required
def add_customer(request, deal_id):
    deal = Deal.objects.get(id=deal_id)
    contacts = Contact.objects.filter(owner=request.user)

    if request.method == 'POST':
        Customer.objects.create(
            contact_id=request.POST['contact'],
            deal=deal,
            lead_status=request.POST['status']
        )
        return redirect('deal_list')

    return render(request, 'deals/add_customer.html', {
        'contacts': contacts,
        'deal': deal
    })

# EDIT DEAL
@login_required
def edit_deal(request, id):
    deal = get_object_or_404(Deal, id=id, owner=request.user)

    if request.method == 'POST':
        deal.title = request.POST['title']
        deal.description = request.POST['description']
        deal.value = request.POST['value']
        deal.expected_close_date = request.POST['expected_close_date']
        deal.save()
        return redirect('deal_list')

    return render(request, 'deals/edit_deal.html', {'deal': deal})


# DELETE DEAL
@login_required
def delete_deal(request, id):
    deal = get_object_or_404(Deal, id=id, owner=request.user)
    deal.delete()
    return redirect('deal_list')

@login_required
def deal_view(request, id):
    deal = Deal.objects.get(id=id)
    customers = Customer.objects.filter(deal=deal)

    return render(request, 'deals/view.html', {
        'deal': deal,
        'customers': customers
    })


# EDIT CUSTOMER
@login_required
def edit_customer(request, id):
    customer = get_object_or_404(Customer, id=id)

    if request.method == 'POST':
        customer.lead_status = request.POST['status']
        customer.save()
        return redirect('deal_list')

    return render(request, 'deals/edit_customer.html', {'customer': customer})


# DELETE CUSTOMER
@login_required
def delete_customer(request, id):
    customer = get_object_or_404(Customer, id=id)
    customer.delete()
    return redirect('deal_list')